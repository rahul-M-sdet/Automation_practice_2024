import openpyxl

file = 'C:\\Users\\RAHUL\\Documents\\Read.xlsx'
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

#for counthe row and col

row = sheet.max_row
column = sheet.max_column

#read the data from row and col

for r in range(1, row+1):
    for c in range(1, column+1):
        print(sheet.cell(r,c).value,  end=' ')
    print()
