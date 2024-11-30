import openpyxl

file = "C:\\Users\\RAHUL\\OneDrive\\ドキュメント\\Data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "Kumar"
sheet.cell(1,3).value = 'xyz'

sheet.cell(2, 1).value = 456
sheet.cell(2, 2).value = "KP"
sheet.cell(2, 3).value = "ABC"

sheet.cell(3, 1).value = 123
sheet.cell(3, 2).value = "Sher"
sheet.cell(3, 3).value = "KP"

workbook.save(file)