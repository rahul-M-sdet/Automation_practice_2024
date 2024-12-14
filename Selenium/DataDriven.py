import openpyxl

# file/workbook/sheet/row/column/loop/
file = "C:\\Users\\RAHUL\\OneDrive\\ドキュメント\\Data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

# find the row and col in excel
rows = sheet.max_row
columns = sheet.max_column

# iterate over the row and col to get the value
for r in range(1, rows + 1):
    for c in range(1, columns + 1):
        print(sheet.cell(r, c).value, end="    ")
    print()
